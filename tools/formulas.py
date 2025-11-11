from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

dachboard_border = Border(left=Side(style='dashed'), right=Side(style='dashed'), top=Side(style='dashed'), bottom=Side(style='dashed'))

none_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

def insertar_formulas_porcentaje(worksheet, column_names, data_values, DATA_START_ROW, FILA_TOTALES_HEADER):
        """
        Inserta fórmulas de Excel en las columnas de porcentaje y suma de MONTO
        """
        from openpyxl.utils import get_column_letter
        
        FILA_SUMA_MONTO = DATA_START_ROW + len(data_values)
        
        # Encontrar las columnas relevantes
        for col_idx, col_name in enumerate(column_names, start=1):
            col_name_str = str(col_name)
            
            # Insertar suma de MONTO
            if 'AUMENTO / DISMINUCION REAL' in col_name_str and col_name_str not in ['FECHA', 'GENERADORA']:
                col_letter = get_column_letter(col_idx)
                
                # Crear fórmula de suma
                formula = f'=SUM({col_letter}{DATA_START_ROW}:{col_letter}{DATA_START_ROW + len(data_values) - 1})'
                
                cell = worksheet.cell(row=FILA_SUMA_MONTO, column=col_idx)
                cell.value = formula
                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = dachboard_border
                cell.number_format = '0'
            
            # Insertar fórmulas de porcentaje
            if '% DEL AUMENTO / DISMINUCION DEL TOTAL' in col_name_str:
                # Buscar la columna MONTO correspondiente (2 posiciones antes)
                monto_col = col_idx - 2
                
                if monto_col > 0 and monto_col <= len(column_names):
                    # Verificar que sea efectivamente una columna MONTO
                    if 'AUMENTO / DISMINUCION REAL' in str(column_names[monto_col - 1]):
                        monto_letter = get_column_letter(monto_col)
                        
                        # Insertar fórmulas para cada fila de datos
                        for i in range(len(data_values)):
                            row_num = DATA_START_ROW + i
                            # La fórmula referencia la celda de suma total en FILA_SUMA_MONTO
                            formula = f'=IF(${monto_letter}${FILA_SUMA_MONTO}=0,0,{monto_letter}{row_num}/${monto_letter}${FILA_SUMA_MONTO})'
                            cell = worksheet.cell(row=row_num, column=col_idx)
                            cell.value = formula
                            cell.number_format = '0%'

            if '% DE AUMENTO / DISMINUCION IDEAL' in col_name_str:
                # Referencias fijas: primera fila = C16, segunda = C17, tercera = C18
                referencias_fijas = ['=$C$16', '=$C$17', '=$C$18']
                
                for i in range(len(data_values)):
                    row_num = DATA_START_ROW + i
                    # Asignar la referencia fija según el índice de la fila
                    if i < len(referencias_fijas):
                        formula = referencias_fijas[i]
                    else:
                        formula = referencias_fijas[-1]  # Si hay más filas, usar la última referencia
                    
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.value = formula
                    cell.number_format = '0%'

            if '% DIFERENCIA' in col_name_str:
                col_del_aumento = col_idx - 2  
                col_ideal = col_idx - 1      
                
                if col_del_aumento > 0 and col_ideal > 0:
                    letra_del_aumento = get_column_letter(col_del_aumento)
                    letra_ideal = get_column_letter(col_ideal)
                    
                    for i in range(len(data_values)):
                        row_num = DATA_START_ROW + i
                        # Fórmula: % DEL AUMENTO - % IDEAL
                        formula = f'={letra_del_aumento}{row_num}-{letra_ideal}{row_num}'
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        cell.value = formula
                        cell.number_format = '0%'

            if 'AUMENTO / DISMINUCION IDEAL' in col_name_str:
                # Buscar la columna MONTO correspondiente (2 posiciones antes)
                monto_col = col_idx - 2
                monto_sum = col_idx - 5
                
                if monto_col > 0 and monto_col <= len(column_names):
                    # Verificar que sea efectivamente una columna MONTO
                    if '% DE AUMENTO / DISMINUCION IDEAL' in str(column_names[monto_col - 1]):

                        monto_letter = get_column_letter(monto_col)
                        monto_sum_letter = get_column_letter(monto_sum)                        
                        
                        # Insertar fórmulas para cada fila de datos
                        for i in range(len(data_values)):
                            row_num = DATA_START_ROW + i
                            # La fórmula referencia la celda de suma total en FILA_SUMA_MONTO
                            formula = f"={monto_letter}{row_num}*${monto_sum_letter}${FILA_SUMA_MONTO}"
                            cell = worksheet.cell(row=row_num, column=col_idx)
                            cell.value = formula
                            cell.number_format = '0.00'


              #============================================              
            if 'AUMENTO / DISMINUCION IDEAL(MW)' in col_name_str and col_name_str not in ['FECHA', 'GENERADORA']:
                col_letter = get_column_letter(col_idx)
                
                # Crear fórmula de suma
                formula = f'=SUM({col_letter}{DATA_START_ROW}:{col_letter}{DATA_START_ROW + len(data_values) - 1})'
                
                cell = worksheet.cell(row=FILA_SUMA_MONTO, column=col_idx)
                cell.value = formula
                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = dachboard_border
                cell.number_format = '0'

              #============================================              
            if 'AUMENTO / DISMINUCION IDEAL(MW)' in col_name_str and col_name_str not in ['FECHA', 'GENERADORA']:
                col_letter = get_column_letter(col_idx)
                
                monto_real_col = col_idx - 5
                
                if monto_real_col > 0 and monto_real_col <= len(column_names):
                    if 'AUMENTO / DISMINUCION REAL' in str(column_names[monto_real_col - 1]):
                        monto_real_letter = get_column_letter(monto_real_col)

                        formula = f'={col_letter}{FILA_SUMA_MONTO - 3}-{monto_real_letter}{FILA_SUMA_MONTO - 3}'
                        
                        cell = worksheet.cell(row=FILA_SUMA_MONTO + 2, column=col_idx)
                        cell.value = formula
                        cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = none_border
                        cell.number_format = '0.00'

                        next_hours = get_column_letter(col_idx + 1)
                        now_hours = get_column_letter(col_idx -6)
                        
                        if col_idx == len(column_names):
                            cell = worksheet.cell(row=FILA_SUMA_MONTO + 3, column=col_idx)
                            cell.value = "12:00:00 AM"
                            cell.number_format = 'mm:ss.0;@'
                        else:
                            formula = f'={next_hours}{FILA_SUMA_MONTO - 5}-{now_hours}{FILA_SUMA_MONTO - 5}'
                            cell = worksheet.cell(row=FILA_SUMA_MONTO + 3, column=col_idx)
                            cell.value = formula
                            cell.number_format = 'mm:ss.0;@'
                        
                        cell = worksheet.cell(row=FILA_SUMA_MONTO + 3, column=col_idx)
                        cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = none_border
                        

                        formula = f'=({col_letter}{FILA_SUMA_MONTO+2}*{col_letter}{FILA_SUMA_MONTO + 3})/60'

                        cell = worksheet.cell(row=FILA_SUMA_MONTO + 4, column=col_idx)
                        cell.value = formula
                        cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = none_border
                        cell.number_format = '0.000000'