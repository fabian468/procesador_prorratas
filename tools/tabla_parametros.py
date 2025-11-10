from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_tabla_parametros(worksheet, DATA_START_ROW, data_values, column_names):
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Estilos
        header_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        white_font = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        
        FILA_INICIO = 10
        
        # Fila 10: Diferencia Real vs Ideal (MWh)
        cell = worksheet.cell(row=FILA_INICIO, column=1)
        cell.value = "Diferencia Real vs Ideal (MWh)"
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Fila 11: Tiempo entre consignas
        cell = worksheet.cell(row=FILA_INICIO + 1, column=1)
        cell.value = "Tiempo entre consignas"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Fila 12: P&L MWh
        cell = worksheet.cell(row=FILA_INICIO + 2, column=1)
        cell.value = "P&L MWh"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        cell = worksheet.cell(row=FILA_INICIO + 2, column=2)
        cell.value = 0
        cell.number_format = '0.000000'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Fila 13: US$ / MWh
        cell = worksheet.cell(row=FILA_INICIO + 3, column=1)
        cell.value = "US$ / MWh"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        cell = worksheet.cell(row=FILA_INICIO + 3, column=2)
        cell.value = 80.00
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formato de dinero en pesos chilenos
        cell.number_format = '"$"#,##0.00'
    
        
        # Fila 14: P&L US$
        cell = worksheet.cell(row=FILA_INICIO + 4, column=1)
        cell.value = "P&L US$"
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        cell = worksheet.cell(row=FILA_INICIO + 4, column=2)
        cell.value = 0.31
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '"$"#,##0.00'
        
        
        # Fila 15: Headers de la tabla verde
        FILA_TABLA = FILA_INICIO + 5
        
        cell = worksheet.cell(row=FILA_TABLA, column=1)
        cell.value = "MWac max de Generación"
        cell.fill = header_green
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        cell = worksheet.cell(row=FILA_TABLA, column=2)
        cell.value = "SITE"
        cell.fill = header_green
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        cell = worksheet.cell(row=FILA_TABLA, column=3)
        cell.value = "% DEL TOTAL"
        cell.fill = header_green
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # Obtener generadoras únicas de los datos
        generadoras = []
        generadoras_nombres = []
        for row in data_values:
            if len(row) > 1 and row[1] not in generadoras_nombres:
                generadoras_nombres.append(row[1])
        
        # Valores de MWac (estos son de ejemplo, ajústalos según tus datos reales)
        mwac_values = {
            'PFV-ELPELICANO': 105,
            'PFV-ELROMERO': 196,
            'PFV-LAHUELLA': 85
        }
        
        total_mwac = sum(mwac_values.values())
        
        # Filas 16-18: Datos de generadoras
        fila_actual = FILA_TABLA + 1
        for gen_nombre in generadoras_nombres:
            # Columna A: MWac
            cell = worksheet.cell(row=fila_actual, column=1)
            cell.value = mwac_values.get(gen_nombre, 0)
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.number_format = '0'
            
            # Columna B: Nombre generadora
            cell = worksheet.cell(row=fila_actual, column=2)
            cell.value = gen_nombre
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
            # Columna C: % del total (fórmula)
            cell = worksheet.cell(row=fila_actual, column=3)
            col_a = get_column_letter(1)
            suma_total_row = FILA_TABLA + len(generadoras_nombres) + 1
            formula = f'=IF(${col_a}${suma_total_row}=0,0,{col_a}{fila_actual}/${col_a}${suma_total_row})'
            cell.value = formula
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.number_format = '0%'
            
            fila_actual += 1
        
        # Fila 19: Suma total
        cell = worksheet.cell(row=fila_actual, column=1)
        formula = f'=SUM(A{FILA_TABLA + 1}:A{fila_actual - 1})'
        cell.value = formula
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.number_format = '0'
        
        # Ajustar anchos de columnas
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 15
    