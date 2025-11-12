from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.utils import get_column_letter

def generarGrafico(column_names, data_values, generadoras, worksheet, fila_actual, DATA_START_ROW):
    try:
        chart = LineChart()
        chart.title = "CONSIGNAS"
        chart.style = 2 
        chart.y_axis.title = "Consigna (MW)"
        chart.x_axis.title = "Horas"
        
    
        chart.smooth = False 

        colores = [
            "4472C4",  # Azul
            "d32213",  # Naranja/Rosa
            "70AD47",  # Verde
            "FFC000",  # Amarillo
        ]

        columnas_consigna = []
        for i, col_name in enumerate(column_names, start=1):
            if 'NUEVO SET POINT POTENCIA ACTIVA(MW)' in str(col_name).upper():
                columnas_consigna.append(i)
        
        if not columnas_consigna:
            print("No se encontraron columnas con 'NUEVO SET POINT' para graficar.")
            return
        
        fila_horas = 3
        columnas_horas = [col - 2 for col in columnas_consigna]
        
        fila_aux_cat = DATA_START_ROW + len(data_values) + 100
        fila_aux_data_start = fila_aux_cat + 1
        col_aux_start = 3
        
        for i, col_hora in enumerate(columnas_horas):
            valor_hora = worksheet.cell(row=fila_horas, column=col_hora).value
            worksheet.cell(row=fila_aux_cat, column=col_aux_start + i, value=valor_hora)
        
        cats = Reference(
            worksheet,
            min_col=col_aux_start,
            min_row=fila_aux_cat,
            max_col=col_aux_start + len(columnas_horas) - 1,
            max_row=fila_aux_cat
        )

        filas_aux_generadoras = []
        for idx, gen in enumerate(generadoras):
            fila_gen = None
            for i, row_data in enumerate(data_values):
                if row_data[1] == gen:
                    fila_gen = DATA_START_ROW + i
                    break
            
            if fila_gen is None:
                print(f"Advertencia: no se encontró fila para la generadora '{gen}'")
                continue

            fila_aux_gen = fila_aux_data_start + len(filas_aux_generadoras)
            filas_aux_generadoras.append((gen, fila_aux_gen))
            
            for i, col_consigna in enumerate(columnas_consigna):
                valor = worksheet.cell(row=fila_gen, column=col_consigna).value
                worksheet.cell(row=fila_aux_gen, column=col_aux_start + i, value=valor)
            
            values = Reference(
                worksheet,
                min_col=col_aux_start,
                min_row=fila_aux_gen,
                max_col=col_aux_start + len(columnas_consigna) - 1,
                max_row=fila_aux_gen
            )
            
            serie = Series(values, title=gen)
            
            color_hex = colores[idx % len(colores)]
            serie.graphicalProperties.line.solidFill = color_hex
            serie.graphicalProperties.line.width = 20000
            
            serie.marker.symbol = "circle" 
            serie.marker.size = 6
            serie.marker.graphicalProperties.solidFill = color_hex
            serie.marker.graphicalProperties.line.solidFill = color_hex
            
            chart.series.append(serie)

        chart.set_categories(cats)

        chart.height = 12
        chart.width = 20
        chart.legend.position = 'b' 
        
        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1

        worksheet.add_chart(chart, "E15")

    except Exception as e:
        print(f"Error al crear el gráfico: {e}")
        import traceback
        traceback.print_exc()